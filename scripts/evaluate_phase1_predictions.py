#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Iterable


STAGES = ["TA43", "TA01", "TA03", "TA11", "TN01"]
TECHNIQUE_TO_STAGE = {
    "TA43_01": "TA43", "TA43_02": "TA43",
    "TA01_01": "TA01", "TA01_02": "TA01",
    "TA03_01": "TA03",
    "TA11_01": "TA11", "TA11_02": "TA11",
    "TN01_01": "TN01",
}
ALIASES = {
    "pcap": ["pcap", "pcapid", "pcap文件", "pcap文件名", "文件名", "抓包文件", "pcapfilename", "pcapfile"],
    "number": ["编号", "number", "序号", "行号", "recordid", "record", "sessionid", "session", "flowid", "flow", "rowid", "row", "id"],
    "stage": ["攻击阶段编号或正常流量编号", "stagecode", "attackstage", "stage", "阶段", "阶段编号", "label"],
    "technique": ["攻击技术编号或正常流量编号", "techniqueguess", "techniquecode", "predictedcode", "攻击技术编号", "技术编号"],
    "start": ["开始时间", "starttime", "start"],
    "end": ["结束时间", "endtime", "end"],
    "src_ip": ["源ip", "srcip", "sourceip"],
    "src_port": ["源端口", "srcport", "sourceport"],
    "dst_ip": ["目的ip", "dstip", "destinationip"],
    "dst_port": ["目的端口", "dstport", "destinationport"],
}


def normalized_header(value: Any) -> str:
    return re.sub(r"[\s_()（）\-:/]+", "", str(value or "").strip().lower())


def column_map(headers: Iterable[Any]) -> dict[str, str]:
    normalized = {normalized_header(header): str(header) for header in headers if header is not None}
    out: dict[str, str] = {}
    for logical, aliases in ALIASES.items():
        for alias in aliases:
            if normalized_header(alias) in normalized:
                out[logical] = normalized[normalized_header(alias)]
                break
    return out


def read_csv_table(path: Path) -> list[dict[str, Any]]:
    last_error: UnicodeDecodeError | None = None
    for encoding in ("utf-8-sig", "utf-8", "gb18030"):
        try:
            with path.open("r", encoding=encoding, newline="") as handle:
                return [dict(row) for row in csv.DictReader(handle)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"cannot decode CSV {path}: {last_error}")


def read_xlsx_table(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("reading .xlsx requires openpyxl; install requirements.txt") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    try:
        headers = [str(value or "") for value in next(rows)]
    except StopIteration:
        return []
    return [dict(zip(headers, values)) for values in rows if any(value not in (None, "") for value in values)]


def read_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return read_csv_table(path)
    if suffix in {".xlsx", ".xlsm"}:
        return read_xlsx_table(path)
    raise ValueError(f"unsupported table type: {path.suffix}; use .csv or .xlsx")


def clean(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return text[:-2] if re.fullmatch(r"\d+\.0", text) else text


def pcap_key(value: Any) -> str:
    return Path(clean(value).replace("\\", "/")).name.lower()


def normalize_stage(value: Any) -> str:
    text = clean(value).upper()
    if text in STAGES:
        return text
    if text in TECHNIQUE_TO_STAGE:
        return TECHNIQUE_TO_STAGE[text]
    compact = normalized_header(value)
    words = {
        "reconnaissance": "TA43", "侦察": "TA43", "扫描": "TA43",
        "initialaccess": "TA01", "初始访问": "TA01", "漏洞利用": "TA01",
        "persistence": "TA03", "持久化": "TA03", "投递": "TA03",
        "commandandcontrol": "TA11", "c2": "TA11", "命令与控制": "TA11", "回连": "TA11",
        "normal": "TN01", "正常": "TN01", "正常流量": "TN01",
    }
    return words.get(compact, "")


def value(row: dict[str, Any], columns: dict[str, str], logical: str) -> str:
    return clean(row.get(columns.get(logical, "")))


def signature(row: dict[str, Any], columns: dict[str, str]) -> tuple[str, ...]:
    return (
        pcap_key(value(row, columns, "pcap")), value(row, columns, "start"), value(row, columns, "end"),
        value(row, columns, "src_ip"), value(row, columns, "src_port"),
        value(row, columns, "dst_ip"), value(row, columns, "dst_port"),
    )


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def metric(tp: int, fp: int, fn: int) -> tuple[float, float, float]:
    precision = tp / (tp + fp) if tp + fp else 0.0
    recall = tp / (tp + fn) if tp + fn else 0.0
    f1 = 2 * precision * recall / (precision + recall) if precision + recall else 0.0
    return precision, recall, f1


def load_techniques(path: Path | None) -> dict[str, str]:
    if not path or not path.exists():
        return {}
    out: dict[str, str] = {}
    for line in path.read_text(encoding="utf-8").splitlines():
        if not line.strip():
            continue
        item = json.loads(line)
        key = clean(item.get("record_id") or item.get("编号"))
        technique = clean(item.get("technique_guess") or item.get("technique_code")).upper()
        if key and technique:
            out[key] = technique
    return out


def evaluate(predictions: Path, answer: Path, output_dir: Path, predictions_jsonl: Path | None = None) -> dict[str, Any]:
    pred_rows = read_table(predictions)
    answer_rows = read_table(answer)
    if not pred_rows:
        raise ValueError("prediction table has no data rows")
    if not answer_rows:
        raise ValueError("answer table has no data rows")
    pred_cols = column_map(pred_rows[0].keys())
    answer_cols = column_map(answer_rows[0].keys())
    if "stage" not in pred_cols:
        raise ValueError("prediction table has no recognized stage column; accepted examples: stage_code, stage, 阶段, 攻击阶段编号或正常流量编号")
    if "stage" not in answer_cols and "technique" not in answer_cols:
        raise ValueError("answer table has no recognized label column; accepted examples: stage_code, stage, technique_code, 阶段, 技术编号")

    pred_by_primary: dict[tuple[str, str], list[int]] = {}
    pred_by_number: dict[str, list[int]] = {}
    pred_by_signature: dict[tuple[str, ...], list[int]] = {}
    for index, row in enumerate(pred_rows):
        pcap = pcap_key(value(row, pred_cols, "pcap"))
        number = value(row, pred_cols, "number")
        if pcap and number:
            pred_by_primary.setdefault((pcap, number), []).append(index)
        if number:
            pred_by_number.setdefault(number, []).append(index)
        sig = signature(row, pred_cols)
        if any(sig[1:]):
            pred_by_signature.setdefault(sig, []).append(index)

    used: set[int] = set()
    matched: list[tuple[int, int, str]] = []
    unmatched: list[dict[str, Any]] = []
    for answer_index, row in enumerate(answer_rows):
        pcap = pcap_key(value(row, answer_cols, "pcap"))
        number = value(row, answer_cols, "number")
        candidates: list[tuple[str, list[int]]] = []
        if pcap and number:
            candidates.append(("pcap+number", pred_by_primary.get((pcap, number), [])))
        if number:
            candidates.append(("number", pred_by_number.get(number, [])))
        sig = signature(row, answer_cols)
        if any(sig[1:]):
            candidates.append(("observable_signature", pred_by_signature.get(sig, [])))
        chosen = None
        method = ""
        for candidate_method, indexes in candidates:
            available = [index for index in indexes if index not in used]
            if len(available) == 1:
                chosen = available[0]
                method = candidate_method
                break
        if chosen is None:
            unmatched.append({"side": "answer", "answer_row": answer_index + 2, "pcap": pcap, "number": number, "reason": "no unique prediction match"})
            continue
        used.add(chosen)
        matched.append((answer_index, chosen, method))
    for pred_index, row in enumerate(pred_rows):
        if pred_index not in used:
            unmatched.append({"side": "prediction", "prediction_row": pred_index + 2, "pcap": pcap_key(value(row, pred_cols, "pcap")), "number": value(row, pred_cols, "number"), "reason": "unused prediction row"})

    confusion = {(expected, predicted): 0 for expected in STAGES for predicted in STAGES}
    errors: list[dict[str, Any]] = []
    technique_predictions = load_techniques(predictions_jsonl)
    technique_total = technique_correct = 0
    match_methods = Counter()
    valid_pairs = 0
    for answer_index, pred_index, method in matched:
        answer_row, pred_row = answer_rows[answer_index], pred_rows[pred_index]
        expected_technique = value(answer_row, answer_cols, "technique").upper()
        expected = normalize_stage(value(answer_row, answer_cols, "stage") or expected_technique)
        predicted = normalize_stage(value(pred_row, pred_cols, "stage"))
        number = value(pred_row, pred_cols, "number")
        if expected not in STAGES or predicted not in STAGES:
            unmatched.append({"side": "pair", "answer_row": answer_index + 2, "prediction_row": pred_index + 2, "number": number, "reason": f"invalid stage expected={expected or '?'} predicted={predicted or '?'}"})
            continue
        valid_pairs += 1
        match_methods[method] += 1
        confusion[(expected, predicted)] += 1
        if expected != predicted:
            errors.append({"answer_row": answer_index + 2, "prediction_row": pred_index + 2, "pcap": pcap_key(value(pred_row, pred_cols, "pcap")), "number": number, "expected_stage": expected, "predicted_stage": predicted})
        if expected_technique in TECHNIQUE_TO_STAGE and number in technique_predictions:
            technique_total += 1
            technique_correct += technique_predictions[number] == expected_technique

    correct = sum(confusion[(stage, stage)] for stage in STAGES)
    overall = correct / valid_pairs if valid_pairs else 0.0
    normal_total = sum(confusion[("TN01", predicted)] for predicted in STAGES)
    attack_total = valid_pairs - normal_total
    normal_correct = confusion[("TN01", "TN01")]
    attack_correct = sum(confusion[(stage, stage)] for stage in STAGES if stage != "TN01")
    normal_vs_attack_correct = sum(
        count for (expected, predicted), count in confusion.items()
        if (expected == "TN01") == (predicted == "TN01")
    )

    class_metrics = []
    for stage in STAGES:
        tp = confusion[(stage, stage)]
        fp = sum(confusion[(other, stage)] for other in STAGES if other != stage)
        fn = sum(confusion[(stage, other)] for other in STAGES if other != stage)
        precision, recall, f1 = metric(tp, fp, fn)
        class_metrics.append({"stage": stage, "support": tp + fn, "precision": precision, "recall": recall, "f1": f1})

    output_dir.mkdir(parents=True, exist_ok=True)
    confusion_rows = [{"expected": expected, **{predicted: confusion[(expected, predicted)] for predicted in STAGES}} for expected in STAGES]
    write_csv(output_dir / "confusion_matrix.csv", ["expected", *STAGES], confusion_rows)
    write_csv(output_dir / "errors.csv", ["answer_row", "prediction_row", "pcap", "number", "expected_stage", "predicted_stage"], errors)
    write_csv(output_dir / "unmatched_rows.csv", ["side", "answer_row", "prediction_row", "pcap", "number", "reason"], unmatched)

    lines = [
        "# Phase-1 Evaluation Report", "",
        f"- Prediction rows: {len(pred_rows)}",
        f"- Answer rows: {len(answer_rows)}",
        f"- Matched valid rows: {valid_pairs}",
        f"- Unmatched or invalid rows: {len(unmatched)}",
        f"- Overall stage accuracy: {overall:.4f}",
        f"- Normal-vs-attack accuracy: {normal_vs_attack_correct / valid_pairs if valid_pairs else 0.0:.4f}",
        f"- Normal accuracy: {normal_correct / normal_total if normal_total else 0.0:.4f} ({normal_correct}/{normal_total})",
        f"- Attack-stage accuracy: {attack_correct / attack_total if attack_total else 0.0:.4f} ({attack_correct}/{attack_total})",
        f"- Match methods: {dict(match_methods)}", "",
        "## Per-stage metrics", "",
        "| Stage | Support | Precision | Recall | F1 |", "|---|---:|---:|---:|---:|",
    ]
    lines.extend(f"| {row['stage']} | {row['support']} | {row['precision']:.4f} | {row['recall']:.4f} | {row['f1']:.4f} |" for row in class_metrics)
    lines.extend(["", "## Technique guess", ""])
    if technique_total:
        lines.append(f"- Optional technique accuracy: {technique_correct / technique_total:.4f} ({technique_correct}/{technique_total})")
    else:
        lines.append("- Optional technique accuracy: not available (no alignable technique labels and guesses).")
    lines.extend(["", "## Artifacts", "", "- `confusion_matrix.csv`", "- `errors.csv`", "- `unmatched_rows.csv`", ""])
    (output_dir / "eval_report.md").write_text("\n".join(lines), encoding="utf-8")
    return {"matched": valid_pairs, "unmatched": len(unmatched), "accuracy": overall, "errors": len(errors)}


def main() -> int:
    parser = argparse.ArgumentParser(description="Evaluate Phase-1 stage predictions against a CSV/XLSX answer table.")
    parser.add_argument("--predictions", type=Path, required=True)
    parser.add_argument("--answer", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--predictions-jsonl", type=Path)
    args = parser.parse_args()
    result = evaluate(args.predictions, args.answer, args.output_dir, args.predictions_jsonl)
    print(json.dumps(result, ensure_ascii=False))
    return 0 if result["matched"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
