#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import ipaddress
import json
import re
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from build_qwen35_session_prompts import STAGE_CODES, TECHNIQUE_CODES, TECHNIQUE_TO_STAGE
from session_card_indicators import redact_sensitive_text


ROOT = Path(__file__).resolve().parents[1]
OFFICIAL_COLUMNS = [
    "pcap编号",
    "开始时间",
    "结束时间",
    "源IP",
    "源端口",
    "目的IP",
    "目的端口",
    "攻击阶段编号或正常流量编号",
    "研判理由（不计分）",
]
DEBUG_REASON_FIELDS = ["研判理由（不计入评分）", "研判理由（不计分）", "reason"]
PCAP_ID_SOURCES = {"pcap_id", "pcap_name", "filename_stem"}
SUBMISSION_LABEL_LEVELS = {"stage", "technique"}
SUBMISSION_TIMEZONES = {"UTC", "Asia/Shanghai"}
OFFICIAL_METADATA_SOURCES = {"representative", "aggregate"}
TEMPLATE_ID_FIELDS = ("pcap编号", "文件名", "pcap", "pcap_id", "pcap_name", "filename", "file_name")
TEMPLATE_METADATA_FIELDS = ["开始时间", "结束时间", "源IP", "源端口", "目的IP", "目的端口"]
PLACEHOLDER_IPS = {"1.1.1.1", "2.2.2.2"}
EPOCH_MIN = 946684800
EPOCH_MAX = 4102444800


def non_empty(value: Any) -> bool:
    return value not in (None, "", [], {}, "-")


def first_value(row: dict[str, Any], *keys: str) -> Any:
    for key in keys:
        value = row.get(key)
        if non_empty(value):
            return value
    return ""


def dict_items(value: Any) -> list[dict[str, Any]]:
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    if isinstance(value, dict):
        return [value]
    return []


def load_json(path: Path, default: Any = None) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


def load_jsonl(path: Path) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    if not path.exists():
        return rows
    with path.open("r", encoding="utf-8") as handle:
        for line in handle:
            line = line.strip()
            if line:
                item = json.loads(line)
                if isinstance(item, dict):
                    rows.append(item)
    return rows


def load_csv(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        return []
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def load_xlsx(path: Path) -> list[dict[str, Any]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("xlsx submission templates require openpyxl") from exc
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "").strip() for value in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        item = {headers[index]: value for index, value in enumerate(raw) if index < len(headers) and headers[index]}
        if any(non_empty(value) for value in item.values()):
            out.append(item)
    return out


def load_table(path: Path) -> list[dict[str, Any]]:
    suffix = path.suffix.lower()
    if suffix in {".xlsx", ".xlsm"}:
        return load_xlsx(path)
    return load_csv(path)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OFFICIAL_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, Any]]) -> bool:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "official_submission"
    sheet.append(OFFICIAL_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    for row in rows:
        sheet.append([row.get(column, "") for column in OFFICIAL_COLUMNS])
    widths = {
        "A": 22,
        "B": 24,
        "C": 24,
        "D": 18,
        "E": 12,
        "F": 18,
        "G": 12,
        "H": 28,
        "I": 72,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width
    workbook.save(path)
    return True


def timezone_for_name(name: str) -> timezone:
    if name == "UTC":
        return timezone.utc
    if name == "Asia/Shanghai":
        return timezone(timedelta(hours=8), "Asia/Shanghai")
    raise ValueError(f"submission_timezone must be one of {sorted(SUBMISSION_TIMEZONES)}")


def format_epoch(value: float, submission_timezone: str) -> str:
    text = datetime.fromtimestamp(value, timezone_for_name(submission_timezone)).strftime("%Y-%m-%d %H:%M:%S.%f")
    return text.rstrip("0").rstrip(".")


def format_time(value: Any, submission_timezone: str = "Asia/Shanghai") -> str:
    if not non_empty(value):
        return ""
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return format_epoch(float(value), submission_timezone) if EPOCH_MIN <= float(value) <= EPOCH_MAX else f"{value:g}"
    text = str(value).strip()
    try:
        number = float(text)
    except ValueError:
        return " ".join(text.split())
    if EPOCH_MIN <= number <= EPOCH_MAX:
        return format_epoch(number, submission_timezone)
    if number.is_integer():
        return str(int(number))
    return f"{number:.6f}".rstrip("0").rstrip(".")


def stage_from_technique(value: Any) -> str:
    technique = str(value or "").strip().upper()
    if technique in TECHNIQUE_TO_STAGE:
        return TECHNIQUE_TO_STAGE[technique]
    prefix = technique.split("_", 1)[0]
    return prefix if prefix in STAGE_CODES else ""


def normalize_stage(value: Any) -> str:
    text = str(value or "").strip().upper()
    if text in STAGE_CODES:
        return text
    return stage_from_technique(text)


def normalize_technique(value: Any) -> str:
    text = str(value or "").strip().upper()
    return text if text in TECHNIQUE_CODES else ""


def label_for_submission(row: dict[str, Any], label_level: str) -> str:
    technique = normalize_technique(first_value(row, "technique_guess", "technique_code", "predicted_code"))
    stage = normalize_stage(first_value(row, "stage_code", "攻击阶段编号或正常流量编号", "stage"))
    if not stage:
        stage = stage_from_technique(technique)
    if label_level == "technique":
        return technique or stage
    return stage or stage_from_technique(technique) or technique


def pcap_number(row: dict[str, Any], source: str) -> str:
    pcap_id = str(first_value(row, "pcap_id", "case_id")).strip()
    pcap_name = str(first_value(row, "pcap_name", "pcap", "filename", "file_name")).strip()
    if source == "pcap_name":
        return pcap_name or pcap_id
    if source == "filename_stem":
        name = pcap_name or pcap_id
        return Path(name).stem if name else ""
    return pcap_id or pcap_name


def scrub_reason(value: Any, max_chars: int) -> str:
    text = " ".join(str(value or "").split())
    text = redact_sensitive_text(text, max_chars)
    text = re.sub(r"(?i)\b(api[_-]?key|secret[_-]?key|authorization)\s*[:=]\s*\S+", r"\1=[REDACTED]", text)
    return text[:max_chars]


def cell_value(value: Any) -> str:
    if not non_empty(value):
        return ""
    if isinstance(value, list):
        return ",".join(str(item) for item in value if non_empty(item))
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return str(value)


def is_placeholder_ip(value: Any) -> bool:
    return str(value or "").strip() in PLACEHOLDER_IPS


def is_ip_like(value: Any) -> bool:
    try:
        ipaddress.ip_address(str(value or "").strip())
        return True
    except ValueError:
        return False


def endpoint_ip_score(value: Any) -> int:
    text = str(value or "").strip()
    if not text or text.lower() == "multiple":
        return -20
    if is_placeholder_ip(text):
        return -120
    if is_ip_like(text):
        return 45
    return 5


def endpoint_port_score(value: Any, *, allow_collection: bool = False) -> int:
    if not non_empty(value):
        return -8
    text = str(value).strip().lower()
    if text == "multiple":
        return 2 if allow_collection else -4
    if isinstance(value, list) or "," in text:
        return 8 if allow_collection else 2
    return 12


def aggregate_metadata(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "start_time": first_value(row, "start_time", "开始时间", "ts"),
        "end_time": first_value(row, "end_time", "结束时间"),
        "src_ip": first_value(row, "src_ip", "源IP"),
        "src_port": first_value(row, "src_port", "源端口"),
        "dst_ip": first_value(row, "dst_ip", "目的IP"),
        "dst_port": first_value(row, "dst_port", "目的端口"),
    }


def candidate_metadata(candidate: dict[str, Any]) -> dict[str, Any]:
    return {
        "start_time": first_value(candidate, "start_time", "scan_start", "ts"),
        "end_time": first_value(candidate, "end_time", "scan_end"),
        "src_ip": first_value(candidate, "src_ip", "源IP"),
        "src_port": first_value(candidate, "src_port", "源端口"),
        "dst_ip": first_value(candidate, "dst_ip", "目的IP"),
        "dst_port": first_value(candidate, "dst_port", "目的端口"),
    }


def merge_metadata(preferred: dict[str, Any], fallback: dict[str, Any]) -> dict[str, Any]:
    return {key: preferred.get(key) if non_empty(preferred.get(key)) else fallback.get(key) for key in fallback}


def candidate_score(candidate: dict[str, Any], *, scan: bool = False) -> int:
    metadata = candidate_metadata(candidate)
    score = 0
    score += endpoint_ip_score(metadata.get("src_ip"))
    score += endpoint_ip_score(metadata.get("dst_ip"))
    score += endpoint_port_score(metadata.get("src_port"))
    score += endpoint_port_score(metadata.get("dst_port"), allow_collection=scan)
    if non_empty(metadata.get("start_time")):
        score += 4
    if non_empty(metadata.get("end_time")):
        score += 4
    if is_placeholder_ip(metadata.get("src_ip")) and is_placeholder_ip(metadata.get("dst_ip")):
        score -= 80
    return score


def select_best_candidate(candidates: list[dict[str, Any]], *, scan: bool = False) -> dict[str, Any]:
    if not candidates:
        return {}
    first = candidates[0]
    best = max(enumerate(candidates), key=lambda item: (candidate_score(item[1], scan=scan), -item[0]))[1]
    return best if candidate_score(best, scan=scan) > candidate_score(first, scan=scan) else first


def compact_ports(values: Any, limit: int = 12) -> str:
    if isinstance(values, list):
        out: list[str] = []
        for item in values:
            value = item.get("value") if isinstance(item, dict) else item
            if non_empty(value):
                out.append(str(value))
            if len(out) >= limit:
                break
        return ",".join(out)
    return cell_value(values)


def scan_port_metadata(row: dict[str, Any], candidate: dict[str, Any]) -> str:
    for key in ("dst_ports_sample", "dst_ports", "ports", "dst_port_set"):
        value = first_value(candidate, key)
        if non_empty(value):
            return compact_ports(value)
    scan_summary = row.get("scan_group_summary") or {}
    for key in ("dst_ports_sample", "dst_ports", "ports", "dst_port_set"):
        value = first_value(scan_summary, key)
        if non_empty(value):
            return compact_ports(value)
    if non_empty(row.get("top_dst_ports")):
        return compact_ports(row.get("top_dst_ports"))
    return cell_value(first_value(candidate, "dst_port"))


def count_value(value: Any) -> int:
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def is_port_scan(row: dict[str, Any], stage: str, technique: str) -> bool:
    scan_summary = row.get("scan_group_summary") or {}
    primary = str(row.get("primary_rule_candidate") or "").upper()
    scan_count = count_value(scan_summary.get("scan_like_record_count"))
    return technique == "TA43_01" or primary == "TA43_01" or (not technique and stage == "TA43" and scan_count > 0)


def representative_metadata(row: dict[str, Any]) -> dict[str, Any]:
    fallback = aggregate_metadata(row)
    technique = normalize_technique(first_value(row, "technique_guess", "technique_code", "predicted_code"))
    stage = normalize_stage(first_value(row, "stage_code", "攻击阶段编号或正常流量编号", "stage")) or stage_from_technique(technique)
    if is_port_scan(row, stage, technique):
        candidate = select_best_candidate(dict_items((row.get("scan_group_summary") or {}).get("top_records")), scan=True)
        if not candidate:
            candidate = select_best_candidate(dict_items(row.get("top_suspicious_sessions")), scan=True)
        preferred = candidate_metadata(candidate)
        ports = scan_port_metadata(row, candidate)
        if non_empty(ports):
            preferred["dst_port"] = ports
        return merge_metadata(preferred, fallback)
    if stage == "TN01":
        candidate = select_best_candidate(dict_items(row.get("representative_benign_context")))
        return merge_metadata(candidate_metadata(candidate), fallback)
    candidate = select_best_candidate(dict_items(row.get("top_suspicious_sessions")))
    return merge_metadata(candidate_metadata(candidate), fallback)


def official_metadata(row: dict[str, Any], source: str) -> dict[str, Any]:
    if source == "aggregate":
        return aggregate_metadata(row)
    if source == "representative":
        return representative_metadata(row)
    raise ValueError(f"official_metadata_source must be one of {sorted(OFFICIAL_METADATA_SOURCES)}")


def record_lookup(records_path: Path | None) -> dict[str, dict[str, Any]]:
    if not records_path or not records_path.exists():
        return {}
    records = load_json(records_path, [])
    if not isinstance(records, list):
        return {}
    out: dict[str, dict[str, Any]] = {}
    for record in records:
        if not isinstance(record, dict):
            continue
        record_id = str(record.get("record_id") or record.get("session_id") or "")
        pcap_id = str(record.get("pcap_id") or "")
        if record_id:
            out[record_id] = record
        if pcap_id:
            out.setdefault(pcap_id, record)
    return out


def parse_summary_lookup(parse_summary_path: Path | None) -> dict[str, str]:
    if not parse_summary_path or not parse_summary_path.exists():
        return {}
    summary = load_json(parse_summary_path, [])
    if not isinstance(summary, list):
        return {}
    out: dict[str, str] = {}
    for item in summary:
        if not isinstance(item, dict):
            continue
        case_id = str(item.get("case_id") or "")
        pcap_path = str(item.get("pcap_path") or item.get("pcap") or "")
        if case_id and pcap_path:
            out[case_id] = Path(pcap_path).name
    return out


def template_key_values(value: Any) -> list[str]:
    text = str(value or "").strip()
    if not text:
        return []
    path = Path(text)
    values = [text, path.name, path.stem]
    return [item.lower() for item in values if item]


def template_lookup_keys(row: dict[str, Any], pcap_id_source: str) -> list[str]:
    values = [
        pcap_number(row, pcap_id_source),
        pcap_number(row, "pcap_id"),
        pcap_number(row, "pcap_name"),
        pcap_number(row, "filename_stem"),
        first_value(row, "pcap_id", "case_id"),
        first_value(row, "pcap_name", "pcap", "filename", "file_name"),
    ]
    keys: list[str] = []
    for value in values:
        keys.extend(template_key_values(value))
    return list(dict.fromkeys(keys))


def template_identity(row: dict[str, Any]) -> str:
    return str(first_value(row, *TEMPLATE_ID_FIELDS)).strip()


def load_submission_template(path: Path | None) -> dict[str, dict[str, Any]]:
    if path is None:
        return {}
    if not path.exists():
        raise FileNotFoundError(f"submission template does not exist: {path}")
    rows = load_table(path)
    if not rows:
        return {}
    headers = set(rows[0])
    if not any(field in headers for field in TEMPLATE_ID_FIELDS) or not all(field in headers for field in TEMPLATE_METADATA_FIELDS):
        return {}
    lookup: dict[str, dict[str, Any]] = {}
    for row in rows:
        identity = template_identity(row)
        if not identity:
            continue
        official_row = {
            "pcap编号": first_value(row, "pcap编号", "文件名", "pcap", "pcap_id", "pcap_name", "filename", "file_name"),
            "开始时间": first_value(row, "开始时间"),
            "结束时间": first_value(row, "结束时间"),
            "源IP": first_value(row, "源IP"),
            "源端口": first_value(row, "源端口"),
            "目的IP": first_value(row, "目的IP"),
            "目的端口": first_value(row, "目的端口"),
        }
        for key in template_key_values(identity):
            lookup.setdefault(key, official_row)
    return lookup


def matching_template_row(row: dict[str, Any], template_lookup: dict[str, dict[str, Any]], pcap_id_source: str) -> dict[str, Any]:
    for key in template_lookup_keys(row, pcap_id_source):
        if key in template_lookup:
            return template_lookup[key]
    return {}


def merge_prediction_context(
    prediction: dict[str, Any],
    records: dict[str, dict[str, Any]],
    pcap_names: dict[str, str],
) -> dict[str, Any]:
    record_id = str(prediction.get("record_id") or prediction.get("编号") or "")
    pcap_id = str(prediction.get("pcap_id") or "")
    base = dict(records.get(record_id) or records.get(pcap_id) or {})
    merged = {**base, **prediction}
    if not non_empty(merged.get("pcap_name")):
        mapped = pcap_names.get(str(merged.get("pcap_id") or ""))
        if mapped:
            merged["pcap_name"] = mapped
            merged.setdefault("pcap", mapped)
    return merged


def build_official_rows(
    predictions: list[dict[str, Any]],
    *,
    records: dict[str, dict[str, Any]] | None = None,
    pcap_names: dict[str, str] | None = None,
    template_lookup: dict[str, dict[str, Any]] | None = None,
    pcap_id_source: str = "pcap_id",
    submission_label_level: str = "stage",
    submission_timezone: str = "Asia/Shanghai",
    official_metadata_source: str = "representative",
    max_reason_chars: int = 300,
) -> list[dict[str, Any]]:
    if pcap_id_source not in PCAP_ID_SOURCES:
        raise ValueError(f"pcap_id_source must be one of {sorted(PCAP_ID_SOURCES)}")
    if submission_label_level not in SUBMISSION_LABEL_LEVELS:
        raise ValueError(f"submission_label_level must be one of {sorted(SUBMISSION_LABEL_LEVELS)}")
    if submission_timezone not in SUBMISSION_TIMEZONES:
        raise ValueError(f"submission_timezone must be one of {sorted(SUBMISSION_TIMEZONES)}")
    if official_metadata_source not in OFFICIAL_METADATA_SOURCES:
        raise ValueError(f"official_metadata_source must be one of {sorted(OFFICIAL_METADATA_SOURCES)}")
    record_map = records or {}
    pcap_map = pcap_names or {}
    templates = template_lookup or {}
    rows: list[dict[str, Any]] = []
    for prediction in predictions:
        row = merge_prediction_context(prediction, record_map, pcap_map)
        reason = first_value(row, *DEBUG_REASON_FIELDS)
        template = matching_template_row(row, templates, pcap_id_source)
        if template:
            rows.append({
                "pcap编号": cell_value(first_value(template, "pcap编号")),
                "开始时间": cell_value(first_value(template, "开始时间")),
                "结束时间": cell_value(first_value(template, "结束时间")),
                "源IP": cell_value(first_value(template, "源IP")),
                "源端口": cell_value(first_value(template, "源端口")),
                "目的IP": cell_value(first_value(template, "目的IP")),
                "目的端口": cell_value(first_value(template, "目的端口")),
                "攻击阶段编号或正常流量编号": label_for_submission(row, submission_label_level),
                "研判理由（不计分）": scrub_reason(reason, max_reason_chars),
            })
            continue
        metadata = official_metadata(row, official_metadata_source)
        rows.append({
            "pcap编号": pcap_number(row, pcap_id_source),
            "开始时间": format_time(metadata.get("start_time"), submission_timezone),
            "结束时间": format_time(metadata.get("end_time"), submission_timezone),
            "源IP": cell_value(metadata.get("src_ip")),
            "源端口": cell_value(metadata.get("src_port")),
            "目的IP": cell_value(metadata.get("dst_ip")),
            "目的端口": cell_value(metadata.get("dst_port")),
            "攻击阶段编号或正常流量编号": label_for_submission(row, submission_label_level),
            "研判理由（不计分）": scrub_reason(reason, max_reason_chars),
        })
    return rows


def default_paths(output_dir: Path | None) -> dict[str, Path | None]:
    if output_dir is None:
        return {
            "predictions_jsonl": None,
            "predictions_csv": None,
            "records_json": None,
            "parse_summary": None,
            "output_csv": None,
            "output_xlsx": None,
        }
    return {
        "predictions_jsonl": output_dir / "predictions.jsonl",
        "predictions_csv": output_dir / "phase1_predictions.csv",
        "records_json": output_dir / "session_cards" / "selected_records.json",
        "parse_summary": output_dir / "parsed" / "parse_all_summary.json",
        "output_csv": output_dir / "official_submission.csv",
        "output_xlsx": output_dir / "official_submission.xlsx",
    }


def load_predictions(predictions_jsonl: Path | None, predictions_csv: Path | None) -> tuple[list[dict[str, Any]], str]:
    jsonl_rows = load_jsonl(predictions_jsonl) if predictions_jsonl else []
    if jsonl_rows:
        return jsonl_rows, str(predictions_jsonl)
    csv_rows = load_csv(predictions_csv) if predictions_csv else []
    if csv_rows:
        return csv_rows, str(predictions_csv)
    if predictions_jsonl and predictions_jsonl.exists():
        return [], str(predictions_jsonl)
    if predictions_csv and predictions_csv.exists():
        return [], str(predictions_csv)
    return [], ""


def export_official_submission(
    *,
    predictions: list[dict[str, Any]] | None = None,
    predictions_jsonl: Path | None = None,
    predictions_csv: Path | None = None,
    records_json: Path | None = None,
    parse_summary: Path | None = None,
    submission_template: Path | None = None,
    output_csv: Path,
    output_xlsx: Path | None = None,
    submission_label_level: str = "stage",
    pcap_id_source: str = "pcap_id",
    submission_timezone: str = "Asia/Shanghai",
    official_metadata_source: str = "representative",
    max_reason_chars: int = 300,
    write_excel: bool = True,
) -> dict[str, Any]:
    source = "in_memory"
    if predictions is None:
        predictions, source = load_predictions(predictions_jsonl, predictions_csv)
    records = record_lookup(records_json)
    pcap_names = parse_summary_lookup(parse_summary)
    template_lookup = load_submission_template(submission_template)
    rows = build_official_rows(
        predictions,
        records=records,
        pcap_names=pcap_names,
        template_lookup=template_lookup,
        pcap_id_source=pcap_id_source,
        submission_label_level=submission_label_level,
        submission_timezone=submission_timezone,
        official_metadata_source=official_metadata_source,
        max_reason_chars=max_reason_chars,
    )
    write_csv(output_csv, rows)
    xlsx_written = False
    if write_excel and output_xlsx is not None:
        xlsx_written = write_xlsx(output_xlsx, rows)
    return {
        "rows": len(rows),
        "source": source,
        "csv": str(output_csv),
        "xlsx": str(output_xlsx) if output_xlsx else "",
        "xlsx_written": xlsx_written,
        "submission_label_level": submission_label_level,
        "pcap_id_source": pcap_id_source,
        "submission_timezone": submission_timezone,
        "official_metadata_source": official_metadata_source,
        "submission_template": str(submission_template) if submission_template else "",
        "submission_template_used": bool(template_lookup),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export strict Phase-1 official submission files.")
    parser.add_argument("--output-dir", type=Path, help="Pipeline output directory; supplies default input/output paths.")
    parser.add_argument("--predictions-jsonl", type=Path)
    parser.add_argument("--predictions-csv", type=Path)
    parser.add_argument("--records-json", type=Path)
    parser.add_argument("--parse-summary", type=Path)
    parser.add_argument("--submission-template", type=Path)
    parser.add_argument("--output-csv", type=Path)
    parser.add_argument("--output-xlsx", type=Path)
    parser.add_argument("--submission-label-level", choices=sorted(SUBMISSION_LABEL_LEVELS), default="stage")
    parser.add_argument("--pcap-id-source", choices=sorted(PCAP_ID_SOURCES), default="pcap_id")
    parser.add_argument("--submission-timezone", choices=sorted(SUBMISSION_TIMEZONES), default="Asia/Shanghai")
    parser.add_argument("--official-metadata-source", choices=sorted(OFFICIAL_METADATA_SOURCES), default="representative")
    parser.add_argument("--max-reason-chars", type=int, default=300)
    parser.add_argument("--no-xlsx", action="store_true", help="Skip XLSX export even if openpyxl is installed.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    defaults = default_paths(args.output_dir.resolve() if args.output_dir else None)
    predictions_jsonl = args.predictions_jsonl or defaults["predictions_jsonl"]
    predictions_csv = args.predictions_csv or defaults["predictions_csv"]
    records_json = args.records_json or defaults["records_json"]
    parse_summary = args.parse_summary or defaults["parse_summary"]
    output_csv = args.output_csv or defaults["output_csv"]
    output_xlsx = args.output_xlsx or defaults["output_xlsx"]
    if output_csv is None:
        raise SystemExit("--output-csv or --output-dir is required")
    stats = export_official_submission(
        predictions_jsonl=predictions_jsonl,
        predictions_csv=predictions_csv,
        records_json=records_json,
        parse_summary=parse_summary,
        submission_template=args.submission_template,
        output_csv=output_csv,
        output_xlsx=output_xlsx,
        submission_label_level=args.submission_label_level,
        pcap_id_source=args.pcap_id_source,
        submission_timezone=args.submission_timezone,
        official_metadata_source=args.official_metadata_source,
        max_reason_chars=args.max_reason_chars,
        write_excel=not args.no_xlsx,
    )
    print(json.dumps(stats, ensure_ascii=False, sort_keys=True))
    if output_xlsx and not stats["xlsx_written"] and not args.no_xlsx:
        print("openpyxl is unavailable; wrote CSV only", file=sys.stderr)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
